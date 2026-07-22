"""XLS and XLSX document extractor."""

from __future__ import annotations

from datetime import date, datetime
from io import BytesIO
from typing import Any

import xlrd
from openpyxl import load_workbook

from tools.document import Document, ExtractorResult
from tools.errors import ExtractionError
from tools.extractor_base import BaseExtractor
from tools.helpers import render_markdown_table


class ExcelExtractor(BaseExtractor):
    def extract(self) -> ExtractorResult:
        if self.context.file_extension == ".xlsx":
            return self._extract_xlsx()
        if self.context.file_extension == ".xls":
            return self._extract_xls()
        raise ExtractionError(f"Unsupported spreadsheet format '{self.context.file_extension}'.")

    def _extract_xlsx(self) -> ExtractorResult:
        workbook = load_workbook(BytesIO(self.context.file_bytes), data_only=True)
        documents: list[Document] = []
        sections: list[str] = []
        try:
            for sheet in workbook.worksheets:
                rows = sheet.iter_rows()
                header_cells = next(rows, None)
                if header_cells is None:
                    continue
                headers = self._headers([cell.value for cell in header_cells])
                table_rows: list[list[str]] = []
                for row_number, cells in enumerate(rows, start=2):
                    if not any(cell.value is not None for cell in cells):
                        continue
                    values: list[str] = []
                    document_values: list[str] = []
                    for header, cell in zip(headers, cells, strict=False):
                        value = self._format_value(cell.value)
                        if cell.hyperlink and value:
                            value = f"[{value}]({cell.hyperlink.target})"
                        values.append(value)
                        if value:
                            document_values.append(f"{header}: {value}")
                    table_rows.append(values)
                    documents.append(
                        Document(
                            page_content="; ".join(document_values),
                            metadata={
                                "source": self.context.file_name,
                                "sheet": sheet.title,
                                "row": row_number - 2,
                            },
                        )
                    )
                if table_rows:
                    sections.append(
                        f"## {sheet.title}\n\n{render_markdown_table(headers, table_rows).rstrip()}"
                    )
        finally:
            workbook.close()
        return ExtractorResult(md_content="\n\n".join(sections), documents=documents)

    def _extract_xls(self) -> ExtractorResult:
        try:
            workbook = xlrd.open_workbook(file_contents=self.context.file_bytes, on_demand=True)
        except xlrd.XLRDError as exc:
            raise ExtractionError(
                f"File '{self.context.file_name}' is not a valid .xls file."
            ) from exc

        documents: list[Document] = []
        sections: list[str] = []
        try:
            for sheet in workbook.sheets():
                if sheet.nrows == 0:
                    continue
                headers = self._headers(sheet.row_values(0))
                table_rows: list[list[str]] = []
                for row_number in range(1, sheet.nrows):
                    cells = sheet.row(row_number)
                    if not any(cell.value not in (None, "") for cell in cells):
                        continue
                    values = [self._format_xls_cell(cell, workbook.datemode) for cell in cells]
                    table_rows.append(values)
                    documents.append(
                        Document(
                            page_content="; ".join(
                                f"{header}: {value}"
                                for header, value in zip(headers, values, strict=False)
                                if value
                            ),
                            metadata={
                                "source": self.context.file_name,
                                "sheet": sheet.name,
                                "row": row_number - 1,
                            },
                        )
                    )
                if table_rows:
                    sections.append(
                        f"## {sheet.name}\n\n{render_markdown_table(headers, table_rows).rstrip()}"
                    )
        finally:
            workbook.release_resources()
        return ExtractorResult(md_content="\n\n".join(sections), documents=documents)

    @staticmethod
    def _headers(values: list[Any]) -> list[str]:
        return [
            ExcelExtractor._format_value(value).strip() or f"column_{index + 1}"
            for index, value in enumerate(values)
        ]

    @staticmethod
    def _format_xls_cell(cell: xlrd.sheet.Cell, datemode: int) -> str:
        if cell.ctype == xlrd.XL_CELL_DATE:
            return xlrd.xldate_as_datetime(cell.value, datemode).isoformat()
        return ExcelExtractor._format_value(cell.value)

    @staticmethod
    def _format_value(value: Any) -> str:
        if value is None:
            return ""
        if isinstance(value, (datetime, date)):
            return value.isoformat()
        if isinstance(value, float) and value.is_integer():
            return str(int(value))
        return str(value)
